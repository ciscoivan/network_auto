import netmiko
import textfsm
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side
import time
import smtplib
from email.mime.text import MIMEText
from email.header import Header
from email.mime.multipart import MIMEMultipart

def device_info():
    dev_ip_list=[]
    dev_name_list=[]
    with open('devices_list.txt') as f:
        for line in f.readlines():
            line_s = line.split()
            device_ip = line_s[0]
            device_name = line_s[1]
            dev_ip_list.append(device_ip)
            dev_name_list.append(device_name)
    return dev_ip_list, dev_name_list

def ssh_disp(dev_ip_list, dev_name_list):
    disp_cpu_result_dict = {}
    for device_ip, device_name in zip(dev_ip_list, dev_name_list):
        connection_info = {
            'device_type': 'huawei',
            'ip': device_ip,
            'username': 'ivan',
            'password': '123.com'
        }
        with netmiko.ConnectHandler(**connection_info) as connect:
            print('已成功登录设备： ' + device_ip)
            disp_cpu_result = connect.send_command('display cpu-usage')
        with open('huawei_display_cpu-usage.template') as t:
            template = textfsm.TextFSM(t)
            disp_cpu_result_textfsm = template.ParseTextToDicts(disp_cpu_result)
            disp_cpu_result_dict[device_name] = disp_cpu_result_textfsm
        print(disp_cpu_result_dict)
    return disp_cpu_result_dict


name = []
def openpyxl_f(res_dict):
    date = time.strftime('%Y-%m-%d')
    #print(len(res_dict))
    row_numbers = [n + 2 for n in range(len(res_dict))]  #0+2  1+2
    #print(row_numbers)

    wb = Workbook()
    cpu_ws = wb.create_sheet("CPU_Usage", 0)

    #先设置表头，列宽，并将表头设置为黄色
    cpu_ws['A1'] = '主机名'
    cpu_ws['B1'] = '5秒钟CPU利用率'
    cpu_ws['C1'] = '1分钟CPU利用率'
    cpu_ws['D1'] = '5分钟CPU利用率'
    cpu_ws.column_dimensions['A'].width = 10
    cpu_ws.column_dimensions['B'].width = 15
    cpu_ws.column_dimensions['C'].width = 15
    cpu_ws.column_dimensions['D'].width = 15


    #将字典里的元素拆分出来，以便后续的利用
    hostname_l = []
    CPU_5s_l = []
    CPU_1m_l = []
    CPU_5m_l = []
    for k, v in res_dict.items():
        hostname_l.append(k)
        CPU_5s_l.append(v[0]['CPU_5s'])
        CPU_1m_l.append(v[0]['CPU_1m'])
        CPU_5m_l.append(v[0]['CPU_5m'])

    #写入表格，row_numbers前面定义过，len为5遍历放到列表里，从2开始因表头是1
    for hostname, row in zip(hostname_l, row_numbers):  #2,1 3,1
        cpu_ws.cell(row=row, column=1, value=hostname)
    for cpu_5s, row in zip(CPU_5s_l, row_numbers):
        cpu_ws.cell(row=row, column=2, value=cpu_5s)
    for cpu_1m, row in zip(CPU_1m_l, row_numbers):
        cpu_ws.cell(row=row, column=3, value=cpu_1m)
    for cpu_5m, row in zip(CPU_5m_l, row_numbers):
        cpu_ws.cell(row=row, column=4, value=cpu_5m)

    c = date+"-华为CPU设备巡检.xlsx"
    name.append(c)
    wb.save(f'{date}-华为CPU设备巡检.xlsx')

    print('\n已输出巡检报告')

def send(file_list):
    mail_host = 'smtp.qq.com' #设置服务器
    mail_user = '513046322'
    mail_password = 'ljmnaaxozrlybhcg'


    sender = '513046322@qq.com'  #发送的游戏
    receivers = ['513046322@qq.com']
    log = "ciscoasdasdasd"

    #message = MIMEText('<p style="color:red;">这是一个测试</P>','html','utf-8')

    message = MIMEMultipart()

    message['From'] = Header(sender)
    message['Subject'] = Header('python测试','utf-8')
    for file_name in file_list:
        print("file_name",file_name)
        attr = MIMEText(open(file_name,'rb').read(),'base64','utf-8')
        attr['Content-Type'] = 'application/octet-stream'  #定义流类型
        #attr['Content-Disposition'] = 'attachment;filename='+file_name
        attr.add_header('Content-Disposition', 'attachment', filename=file_name)
        #print(message.as_string())
        message.attach(attr)
    message.attach(MIMEText('网络巡检收集日志','plain','utf-8'))

    try:
        smtpobj = smtplib.SMTP()
        smtpobj.connect(mail_host, 25)
        smtpobj.login(mail_user,mail_password)
        smtpobj.sendmail(sender,receivers,message.as_string())
    except smtplib.SMTPException as e:
        print('error:%s' % e)


def main():
    dev_ip_list, dev_name_list = device_info()
    res_dict = ssh_disp(dev_ip_list, dev_name_list)
    openpyxl_f(res_dict)
    print(name)
    send(file_list=name)


if __name__=='__main__':
    main()
