import textfsm
import re
from tabulate import tabulate
from openpyxl.styles import PatternFill, Border, Side
import netmiko
from openpyxl import Workbook
from  netmiko import ConnectHandler


wb = Workbook()
ws = wb.active
ws.title = '设备板卡端口统计'
ws['A1'] = '千兆口|万兆口'
ws['B1'] = 'up|down'
ws['C1'] = 'input利用率'
ws['D1'] = 'output利用率'

#匹配接口信息
with open('interface') as f , open('inferface2.template') as c:
    fsm = textfsm.TextFSM(c)
    heder = fsm.header
    result = fsm.ParseText(f.read())
    print(result)
#生产接口字典
port_list = []
for port in result:
    portinfo = {
        'portname':'',
        'portlink':'',
        'traff-in':'',
        'traff_out':'',
    }
    portinfo['portname']=port[0]
    portinfo['portlink']=port[1]
    portinfo['traff-in'] = port[2]
    portinfo['traff_out'] = port[3]
    port_list.append(portinfo)
print(port_list)

c = len(port_list)
print(c)
#for循环+if条件判断语句用来生产表格内容

for dev, a in zip(port_list,range(1,c)):

    h = a + 1               #表格行的初始信息
    l = 1                  #表格列的初始信息

    ws.cell(row=h, column=l).value = dev['portname']
    ws.cell(row=h, column=l + 1).value = dev['portlink']
    ws.cell(row=h, column=l + 2).value = dev['traff-in']
    ws.cell(row=h, column=l + 3).value = dev['traff_out']


wb.save('设备板卡端口统计.xlsx')



