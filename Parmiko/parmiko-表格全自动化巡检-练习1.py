import datetime
import threading
import paramiko
import time
import  glob
from queue import Queue

from openpyxl import load_workbook

def exec():
 date2 = time.strftime("%Y-%m-%d",time.localtime())
 wb = load_workbook('simple_excel.xlsx', read_only=True)
 ws1 = wb[wb.sheetnames[0]]
 a = []


 for n in ws1.iter_rows(min_row=2,  min_col=1, ):
    sep = {
        "device_type": n[0].value,
        "ip": n[1].value,
        "username": n[2].value,
        "password": n[3].value,
    }
    a.append(sep)
 return a

def get_command():
     wb = load_workbook('simple_excel.xlsx', read_only=True)
     ws2 = wb[wb.sheetnames[1]]
     command_list = []
     for col in ws2['A2:A5']:
         for cell in col:
             command_list.append(cell.value)
     return command_list

if __name__ == "__main__":
    device = exec()
    command = get_command()
    print(device)
    for i in device:
        print(i['ip'],i['username'],i['password'])
    for i in command:
        print(i)

