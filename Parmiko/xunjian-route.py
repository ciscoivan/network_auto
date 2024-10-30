import paramiko
import time
import os
import shutil
from openpyxl  import  load_workbook
# 指定设备列表、命令列表
wb1 = load_workbook("host-list.xlsx")
wb2 = load_workbook("command.xlsx")
#指定sheet
ws1 =wb1['host']
ws2 =wb2['command']



# 按设备循环
for row1 in range(2 ,ws1.max_row+1):
    try:
        # 获取每行的信息值
        device = ws1.cell(row1, 1).value
        ip = ws1.cell(row1, 2).value
        port = ws1.cell(row1, 3).value
        username = ws1.cell(row1, 4).value
        password = ws1.cell(row1, 5).value
        # 启动ssh连接
        ssh_client = paramiko.SSHClient()
        ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        dev=ssh_client.connect(ip, port=port, username=username, password=password, look_for_keys=False, allow_agent=False)
        print('Successfully connected to',str(ip))
        remote_connection = ssh_client.invoke_shell(width=300)
        remote_connection.send('\n')
        # 循环发送命令
        for row2 in range(2, ws2.max_row + 1):
            command = ws2.cell(row2, 1).value
            remote_connection.send(command+'\n')

        time.sleep(5)
        output = remote_connection.recv(10000000).decode('utf-8').replace('\r','')
        uuid_str = 'R_core_'+device+'_'+ip+'_'+time.strftime("%Y-%m-%d",time.localtime())
        file_name ='%s.txt' % uuid_str
        with open(file_name, 'a+') as f:
            f.write(output)
        ssh_client.close()
        # 将文件移到当天的文件夹下        
        file_path = './%s' % (time.strftime("%Y%m%d", time.localtime()))
        if not os.path.exists(file_path):
            os.makedirs(file_path)
        shutil.move(file_name, '%s/%s' % (file_path, file_name))  # 移动
        print('move file => %s' % (file_name))
    except Exception as e:
        with open("failed_host.txt", 'a+') as f:
            f.write(ip+"\r")
        print(e)
        pass
        



