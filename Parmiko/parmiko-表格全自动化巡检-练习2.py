import datetime
import threading
import paramiko
import time
import glob
from queue import Queue

from openpyxl import load_workbook


def exec():

    wb = load_workbook('simple_excel.xlsx', read_only=True)
    ws1 = wb[wb.sheetnames[0]]
    a = []
    command_list = []

    ws2 = wb[wb.sheetnames[1]]

    for col in ws2['A2:A5']:
        for cell in col:
            command_list.append(cell.value)

    i = 0
    for n in ws1.iter_rows(min_row=2, min_col=1, ):
        sep = {
            "device_type": n[0].value,
            "ip": n[1].value,
            "username": n[2].value,
            "password": n[3].value,

        }
        i = i + 1
        a.append(sep)
    print(a)
    return {
        'device': a, 'command': command_list
    }
if __name__ == "__main__":
    deveice = exec()
    print(deveice)
    m = 0
    for i in deveice:
        print(deveice['device'][m]['ip'])
        print(deveice['device'][m]['username'])
        print(deveice['device'][m]['password'])

        m = m + 1
        for i in deveice['command']:
            print(i)