import datetime
import threading
import paramiko
import time
import glob
from queue import Queue
import smtplib
from email.mime.text import MIMEText
from email.header import Header
from email.mime.multipart import MIMEMultipart
from openpyxl import load_workbook


def exec():

    wb = load_workbook('simple_excel.xlsx', read_only=True)
    ws1 = wb[wb.sheetnames[0]]
    a = []
    command_list = []

    ws2 = wb[wb.sheetnames[1]]

    for col in ws2['A2:A5']:
        for cell in col:
            command_list.append(cell.value)

    i = 0
    for n in ws1.iter_rows(min_row=2, min_col=1, ):
        sep = {
            "device_type": n[0].value,
            "ip": n[1].value,
            "username": n[2].value,
            "password": n[3].value,

        }
        i = i + 1
        a.append(sep)
    return {
        'device': a, 'command': command_list
    }

name = []
def ssh_session(a, command_list):


    cont = 0
    date2 = time.strftime("%Y-%m-%d", time.localtime())
    for da in a:
        file_name = da['ip'] + "-"  + '-info.txt'
        name.append(file_name)
        conn = paramiko.SSHClient()
        conn.set_missing_host_key_policy(paramiko.AutoAddPolicy)
        conn.connect(hostname=da['ip'], username=da['username'], password=da['password'], look_for_keys=False,
                     allow_agent=False)
        with conn.invoke_shell() as conn_new:
            print('*************成功登陆到{}***************'.format(da['ip']))
            conn_new.send("enable\n")
            conn_new.send("123.com\n")
            conn_new.send('term len 0\n')
            for cmd in command_list:
                print('执行巡检命令{}成功！'.format(cmd))
                conn_new.send(cmd +'\n')
            time.sleep(5)
            cont = cont + 1
            output = conn_new.recv(65535).decode()

            with open(file_name, 'a+') as f:
                f.write(output)


    print("****************************************************************")
    print("成功执行{}台设备巡检文件保存在{}目录下!".format(cont, glob.os.getcwd()))




def send(file_list):
    mail_host = 'smtp.qq.com' #设置服务器
    mail_user = '513046322'
    mail_password = 'ljmnaaxozrlybhcg'


    sender = '513046322@qq.com'  #发送的游戏
    receivers = ['513046322@qq.com']
    log = "ciscoasdasdasd"

    #message = MIMEText('<p style="color:red;">这是一个测试</P>','html','utf-8')

    message = MIMEMultipart()

    message['From'] = Header(sender)
    message['Subject'] = Header('python测试','utf-8')
    for file_name in file_list:
        print("file_name",file_name)
        attr = MIMEText(open(file_name,'rb').read(),'base64','utf-8')
        attr['Content-Type'] = 'application/octet-stream'  #定义流类型
        attr['Content-Disposition'] = 'attachment;filename='+file_name
        #print(message.as_string())
        message.attach(attr)
    message.attach(MIMEText('网络巡检收集日志','plain','utf-8'))

    try:
        smtpobj = smtplib.SMTP()
        smtpobj.connect(mail_host, 25)
        smtpobj.login(mail_user,mail_password)
        smtpobj.sendmail(sender,receivers,message.as_string())
    except smtplib.SMTPException as e:
        print('error:%s' % e)

if __name__ == "__main__":
    device = exec()
    ssh = ssh_session(device['device'], device['command'])
    email = send(file_list=name)
