import paramiko
import time
import  glob


from openpyxl import load_workbook

date2 = time.strftime("%Y-%m-%d",time.localtime())


wb = load_workbook('simple_excel.xlsx', read_only=True)
ws1 = wb[wb.sheetnames[0]]

a = []
command_list = []
# 循环遍历行
for n in ws1.iter_rows(min_row=2,  min_col=1, ):
    sep = {
        "device_type": n[0].value,
        "ip": n[1].value,
        "username": n[2].value,
        "password": n[3].value,
    }
    a.append(sep)


print("设备账户登陆账号信息:")
print(a)
print("当前日期为{}:".format(date2))
ws2 = wb[wb.sheetnames[1]]
command_list = []
for col in ws2['A2:A5']:
    for cell in col:
        command_list.append(cell.value)


cont = 0
for da in a:
     file_name =  da['ip'] +"-" +date2+ '-info.txt'
     conn = paramiko.SSHClient()
     conn.set_missing_host_key_policy(paramiko.AutoAddPolicy)
     conn.connect(hostname=da['ip'], username=da['username'], password=da['password'], look_for_keys=False, allow_agent=False)
     with conn.invoke_shell() as conn_new:
         print('*************成功登陆到{}***************'.format(da['ip']))
         conn_new.send("enable\n")
         conn_new.send("123.com\n")
         conn_new.send('term len 0\n')
         for cmd in command_list:
             print('执行巡检命令{}成功！'.format(cmd))
             conn_new.send(cmd + '\n')
         time.sleep(5)
         cont = cont+1
         output = conn_new.recv(65535).decode()

         with open(file_name, 'a+') as f:
             f.write(output)

print("****************************************************************")
print("成功执行{}台设备巡检文件保存在{}目录下!".format(cont,glob.os.getcwd()))
