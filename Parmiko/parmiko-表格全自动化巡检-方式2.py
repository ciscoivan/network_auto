import paramiko
import time

from openpyxl import load_workbook


wb = load_workbook('simple_excel.xlsx', read_only=True)
ws1 = wb[wb.sheetnames[0]]

a = []
# ?????
for n in ws1.iter_rows(min_row=2,  min_col=1, ):
    sep = {
        "device_type": n[0].value,
        "ip": n[1].value,
        "username": n[2].value,
        "password": n[3].value,
    }
    a.append(sep)


wb.close()

print(a)

comands = ['show run\n','show ip int br\n','show ver\n']

for da in a:
     file_name =  da['ip']+'-info.txt'
     print(file_name)
     conn = paramiko.SSHClient()
     conn.set_missing_host_key_policy(paramiko.AutoAddPolicy)
     conn.connect(hostname=da['ip'], username=da['username'], password=da['password'], look_for_keys=False, allow_agent=False)
     with conn.invoke_shell() as conn_new:
         conn_new.send("enable\n")
         conn_new.send("123.com\n")
         conn_new.send('term len 0\n')
         for cmd in comands:
             conn_new.send(cmd + '\n')
         time.sleep(5)

         output = conn_new.recv(65535).decode()

         with open(file_name, 'a+') as f:
             f.write(output)




