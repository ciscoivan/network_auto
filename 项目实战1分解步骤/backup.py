from openpyxl import load_workbook
import time
import  datetime
import os
from netmiko import ConnectHandler
from netmiko.ssh_exception import (NetmikoTimeoutException, AuthenticationException, SSHException, )
from multiprocessing.pool import ThreadPool


class BackupConfig(object):
    def __init__(self):
        self.device_file = "dev-info.xlsx"
        self.pool = ThreadPool(2)
        self.log = os.mkdir('LOG') if os.path.exists('LOG') == False else 'LOG'
        self.logtime = datetime.datetime.now().strftime("%Y-%m-%d_%H_%M_%S")


    def write_to_file(self, *args, **kwargs):
        if kwargs['state'] == 1:
            # 将正常文件写入
            with open(kwargs['path'], 'a') as f:
                f.write(kwargs['result'])
        elif kwargs['state'] == 2:
            # 将测试结果写入文件
            with open(os.path.join(self.log, f'connect_t_{self.logtime}.log'), 'a') as f:
                f.write(kwargs['result'])
                f.write('\n')
        else:
            # 异常文件写入
            with open(os.path.join(self.log, f'error_{self.logtime}.log'), 'a') as f:
                f.write(kwargs['error'])
                f.write('\n')


    def load_execl(self):
        try:
            wb = load_workbook(self.device_file)
            return wb

        except FileNotFoundError:
            print("{}文件不存在".format(self.device_file))


    def get_devices_info(self):
        info_dict = {}
        try:
            wb = self.load_execl()
            ws1 = wb[wb.sheetnames[0]]  # 拿到第一个sheet
            for row in ws1.iter_rows(min_row=2, max_col=9):
                if str(row[1].value).strip() == '#':
                    # 跳过注释行里面带#的
                    continue
                info_dict = {'ip': row[2].value,
                             'protocol': row[3].value,
                             'port': row[4].value,
                             'username': row[5].value,
                             'password': row[6].value,
                             'secret': row[7].value,
                             'device_type': row[8].value,
                             'cmd_list': self.get_cmd_info(wb[row[8].value.strip().lower()]),  # 传参第二个sheet名字
                             }
                yield info_dict
        except Exception as e:
            print('errir', e)

        finally:
            wb.close()


    def get_cmd_info(self, cmd_sheet):
        cmd_list = []
        try:
            for row in cmd_sheet.iter_rows(min_row=2, max_col=2):
                if str(row[0].value).strip() != "#" and row[1].value:
                    # 跳过注释行，去掉命令左右空格
                    cmd_list.append(row[1].value.strip())

            return cmd_list


        except Exception as e:
            print('get_cmd_error', e)


    def connectHandler(self, host):
        try:
            connect = ''
            if host['protocol'].lower().strip() == 'ssh':
                host['port'] = host['port'] if (host['port'] not in [22, None]) else 22
                host.pop('protocol'), host.pop('cmd_list')

                if 'huawei' in host['device_type']:
                    connect = ConnectHandler(**host, conn_timeout=15)

                else:
                    connect = ConnectHandler(**host)

            elif host['protocol'].lower().strip() == 'telnet':
                 host['port'] = host['port'] if (host['port'] not in [23, None]) else 23
                 host.pop('protocol'), host.pop('cmd_list')
                 host['device_type'] = host['device_type'] + '_telnet'
                 connect = ConnectHandler(**host, fast_cli=False)
            else:
                res = "{}_不支持{}协议".format(host['ip'], host['protocol'])
                raise ValueError(res)

            return connect

        except NetmikoTimeoutException as e:
            res = "failid{}连通性问题".format(host['ip'])
            print(res)
            self.write_to_file(**{'state':0,'error':str(e)})
        except AuthenticationException as e:
            res = "{}认证失败".format(host['ip'])
            print(res)
            self.write_to_file(**{'state': 0, 'error': str(e)})
        except SSHException as e:
            res = "{}SSH版本不兼容".format(host['ip'])
            self.write_to_file(**{'state': 0, 'error': str(e)})
        except Exception as e:
            print("connethanler Failed:{}".format(e))
            self.write_to_file(**{'state': 0, 'error': str(e)})


    def run_cmd(self, host, cmds, enable=False):

        enable = True if host['secret'] else False


        conn = self.connectHandler(host)

        if conn:

            hostname = conn.find_prompt().replace('#', '')
            print(hostname)
            # 逐级创建目录
            dirname = host['ip']+'_'+hostname
            dirpath = os.path.join(self.log,self.logtime,dirname)
            print(dirpath)

            try:
                os.makedirs(dirpath)
            except:
                raise Exception("文件件创建失败")
            try:
                if cmds:

                    for cmd in cmds:

                        if enable:
                            conn.enable()
                            print("成功登陆到{}并且执行巡检命令".format(host['ip']))
                            output = conn.send_command(cmd, strip_command=False, strip_prompt=False)
                            data = {'state':1 ,'result': output,'path':os.path.join(dirpath,cmd +'.conf')}
                            self.write_to_file(**data)
                        else:
                            output = conn.send_command(cmd, strip_command=False, strip_prompt=False)
                            data = {'state': 1, 'result': output, 'path': os.path.join(dirpath, cmd + '.conf')}
                            self.write_to_file(**data)
                else:
                          pass
            except Exception as e:
                    print(f"run_cmd Faild:{e}")

            finally:
                conn.close()


    def run_t(self, host):
        # 此模块是测试连接的

        try:
            conn = self.connectHandler(host)
            if conn:
                hostname = conn.find_prompt()
                #print("获取设备的提示符:{}".format(hostname))
                output = "获取设备的提示符:{}".format(hostname)
                self.write_to_file(**({'state': 2, 'result':output}))
                conn.disconnect()

        except Exception as e:
            print(f"run_cmd faild:{e}")


    def connect(self):
        hosts = self.get_devices_info()
        start_time = datetime.datetime.now()
        for host in hosts:
            # self.run_cmd(host,host['cmd_list'])
            # print(host)
            # print(host['cmd_list'])
            # self.pool.apply_async(self.run_t, args=(host, ))
            self.pool.apply_async(self.run_cmd, args=(host, host['cmd_list']))
        self.pool.close()
        self.pool.join()

        end_time = datetime.datetime.now()
        print("-" * 50)
        print("所有任务执行完成，总用时为{:0.2f}秒<<<".format((end_time - start_time).total_seconds()))
        print("-" * 50)

if __name__ == '__main__':
    BackupConfig().connect()
