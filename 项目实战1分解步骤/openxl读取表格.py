import netmiko
import os
import  pandas as pd
from  openpyxl import load_workbook

class BackupConfig(object):
    def __init__(self):
        self.device_file = "dev-info.xlsx"

    def load_execl(self):
        try:
            wb = load_workbook(self.device_file)
            return wb

        except FileNotFoundError:
            print("{}文件不存在".format(self.device_file))

    def get_devices_info(self):
        info_dict = {}
        try:
            wb = self.load_execl()
            ws1 = wb[wb.sheetnames[0]] #拿到第一个sheet
            for row in ws1.iter_rows(min_row=2,max_col=9):
                if str(row[1].value).strip()== '#':
                    #跳过注释行里面带#的
                    continue
                info_dict = {'ip':row[2].value,
                            'protocol':row[3].value,
                            'port':row[4].value,
                            'username':row[5].value,
                            'password':row[6].value,
                            'secret':row[7].value,
                            'device_type':row[8].value,
                            'cmd_list': self.get_cmd_info(wb[row[8].value.strip().lower()]),#传参第二个sheet名字
                             }
                yield info_dict
        except Exception as e:
            print('errir',e)

        finally:
            wb.close()

    def get_cmd_info(self,cmd_sheet):
        cmd_list = []
        try:
            for row in cmd_sheet.iter_rows(min_row=2,max_col=2):
                if str(row[0].value).strip() != "#" and row[1].value:
                    #跳过注释行，去掉命令左右空格
                    cmd_list.append(row[1].value.strip())

            return cmd_list


        except Exception as e:
            print('get_cmd_error',e)


    def connect(self):
        hosts  = self.get_devices_info()
        for host in hosts:
            print(host)

if __name__ == '__main__':
    BackupConfig().connect()
